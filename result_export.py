"""
Export search results to Excel (v6 Phase 5).

Additive layer. Does NOT modify excel_writer, the extraction engine, or the
Phase 1-3 search modules, and never mutates the dataset.

Key guarantees:
- Exports ONLY the SearchResults handed to it (the current/selected results),
  never the whole dataset unless the caller passes all of them.
- Preserves each record's ORIGINAL Published/Granted schema exactly (raw keys,
  no canonical renaming). Mixed datasets are written as one sheet per type so
  each schema stays intact.
- Excel sanitisation stays the single source of truth: the homogeneous case
  calls excel_writer.save_to_excel() directly; the mixed (multi-sheet) case
  reuses the same excel_writer._sanitize() function.
- Optional additive columns: Match, Relevance, Category (never overwriting any
  original field).
- Source PDF cells get a clickable hyperlink to the absolute local path WHEN
  available, while the visible cell keeps the basename (portability preserved).
"""

import os

import pandas as pd
import openpyxl

from excel_writer import save_to_excel, _sanitize

_SOURCE_COLUMN = "Source PDF"


# ----------------------------------------------------------------- row building

def _row_for(result, include_match, include_relevance, categorizer):
    """A COPY of the record's original-schema dict plus optional columns.
    The original record.raw is never modified."""
    row = dict(result.record.raw)          # copy - original schema, no mutation
    if include_match:
        row["Match"] = result.match_summary
    if include_relevance:
        row["Relevance"] = result.relevance
    if categorizer is not None:
        cats = categorizer(result.record)
        row["Category"] = "; ".join(cats) if cats else ""
    return row


def _hyperlink_sheet(worksheet, records):
    """Add a file hyperlink to each Source PDF cell where a source path is
    known; the visible basename in the cell is left unchanged."""
    header = {cell.value: cell.column for cell in worksheet[1]}
    col = header.get(_SOURCE_COLUMN)
    if not col:
        return
    for offset, record in enumerate(records):
        path = getattr(record, "source_path", "")
        if not path:
            continue
        cell = worksheet.cell(row=offset + 2, column=col)
        try:
            cell.hyperlink = os.path.abspath(path)
            cell.style = "Hyperlink"
        except Exception:
            # Never let a link problem break the export.
            pass


def export_results(results, output_file,
                   include_match=True, include_relevance=True, categorizer=None):
    """Export the given SearchResults to `output_file`. Returns the count.

    Homogeneous results -> single sheet via excel_writer.save_to_excel().
    Mixed results -> one sheet per patent type, each in its original schema,
    sanitised with the same excel_writer._sanitize().
    """
    if not results:
        raise ValueError("No results to export.")

    # Group by patent type, remembering both the export row and the record
    # (records are needed to place Source PDF hyperlinks).
    groups = {}
    for result in results:
        row = _row_for(result, include_match, include_relevance, categorizer)
        groups.setdefault(result.record.type, []).append((result.record, row))

    if len(groups) == 1:
        # Homogeneous: literal reuse of the existing sanitising writer.
        records = [rec for rec, _ in next(iter(groups.values()))]
        rows = [row for _, row in next(iter(groups.values()))]
        save_to_excel(rows, output_file)
        workbook = openpyxl.load_workbook(output_file)
        _hyperlink_sheet(workbook.active, records)
        workbook.save(output_file)
    else:
        _write_multi_sheet(groups, output_file)

    return len(results)


def _write_multi_sheet(groups, output_file):
    """One sheet per type, each in its original schema; cells sanitised via the
    same excel_writer._sanitize() used by save_to_excel."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for ptype, items in groups.items():
            rows = [
                {key: _sanitize(value) for key, value in row.items()}
                for _, row in items
            ]
            pd.DataFrame(rows).to_excel(writer, sheet_name=ptype[:31], index=False)

    workbook = openpyxl.load_workbook(output_file)
    for ptype, items in groups.items():
        _hyperlink_sheet(workbook[ptype[:31]], [rec for rec, _ in items])
    workbook.save(output_file)


# ------------------------------------------------------------- open source PDF

def _open_default(path):
    """Open a file with the OS default application."""
    if hasattr(os, "startfile"):        # Windows
        os.startfile(path)              # noqa: S606
        return
    import subprocess
    import sys
    opener = "open" if sys.platform == "darwin" else "xdg-open"
    subprocess.Popen([opener, path])


def open_source_pdf(source_path, page=None, opener=None):
    """Open the source PDF for a record.

    Reliable cross-viewer auto-navigation to a specific page is not available
    from Python, so the PDF is opened normally in the default viewer and the
    page number is returned for the UI to surface (e.g. "opened - see page N").
    Raises ValueError when no path is recorded and FileNotFoundError when the
    file is missing, so the UI can fail gracefully instead of crashing.

    `opener` is injectable for testing.
    """
    if not source_path:
        raise ValueError("No source PDF path is recorded for this record.")
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Source PDF not found on disk:\n{source_path}")

    (opener or _open_default)(source_path)
    return page
